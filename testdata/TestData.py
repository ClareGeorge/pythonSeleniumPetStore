import os

import openpyxl


class TestData:

    @staticmethod
    def getTestData(testcase_name):
        os.chdir("C:\\2.Code\\PythonSelenium\\pythonProjectPetStore\\testdata\\")
        workbook = openpyxl.load_workbook("PetStoreData.xlsx")
        testdata_sheet = workbook["Sheet1"]
        testdata_list = []

        for row in range(1, testdata_sheet.max_row+1):
            if testdata_sheet.cell(row=row, column=1).value == testcase_name:
                testdata_dict = {}
                for col in range(2, testdata_sheet.max_column+1):
                    testdata_dict[testdata_sheet.cell(row=1, column=col).value] = \
                        testdata_sheet.cell(row=row, column=col).value
                testdata_list.append(testdata_dict)
        return testdata_list


